import streamlit as st
import pandas as pd
import io
import csv
import re
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from openpyxl import load_workbook

st.set_page_config(page_title="CTI Sheet -> Meta Import Files", page_icon="📄", layout="centered")
st.title("📄 CTI Sheet -> Meta Import Files")
st.write("Upload your CTI Excel file and generate a single Meta Import CSV (Step1 + Step2).")

uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

# ---------- Helpers ----------

def _norm(name: str) -> str:
    return "".join(name.lower().split())

def get_sheet(wb, wanted: str):
    wanted_norm = _norm(wanted)
    for name in wb.sheetnames:
        if _norm(name) == wanted_norm:
            return wb[name]
    raise KeyError(f"Worksheet '{wanted}' not found. Available: {wb.sheetnames}")

def convert_template(template_name, region):
    if pd.isna(template_name):
        return ""
    mapping = {
        "UCaaS|Link Basic Auto-Attendant": "_AA_Easy",
        "UCaaS|Link Premium Auto-Attendant": "_AA_Premium",
        "UCaaS|Link Lite": "_Lite",
        "UCaaS|Link Standard": "_STD",
        "UCaaS|Link Complete": "_Complete",
        "UCaaS|Link Complete (HIPPA)": "Complete_HIPPA",
        "UCaaS|Link Complete (No Voicemail)": "_Complete_NoVM",
        "UCaaS|Link Complete ContactCenter Agent": "_Complete",
        "UCaaS|Link Complete ContactCenter Manager": "_Complete",
    }
    s = str(template_name).strip()
    return f"{region}{mapping[s]}" if s in mapping else ""

def mac_trusted_until_str():
    """4 weeks out at 11:59:59 pm, formatted m/d/yy h:mm:ss am (no leading apostrophe)."""
    dt = (datetime.now() + timedelta(weeks=4)).replace(hour=23, minute=59, second=59, microsecond=0)
    h12 = dt.hour % 12 or 12
    ampm = "am" if dt.hour < 12 else "pm"
    yy = dt.strftime("%y")
    return f"{dt.month}/{dt.day}/{yy} {h12}:{dt.minute:02d}:{dt.second:02d} {ampm}"

def only_digits(s) -> str:
    return re.sub(r"\D", "", str(s or ""))

def npanxx_from_number(phone: str) -> tuple[str | None, str | None, str | None]:
    """Return (npa, nxx, npanxx) from a phone string; strips leading 1 if present."""
    d = only_digits(phone)
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    if len(d) >= 6:
        npa, nxx = d[:3], d[3:6]
        return npa, nxx, npa + nxx
    return None, None, None

@st.cache_data(ttl=24*3600, show_spinner=False)
def lookup_rc_lata(npa: str, nxx: str) -> tuple[str | None, str | None]:
    """
    NPA-NXX → (Rate Centre, LATA) using LocalCallingGuide XML endpoint.
    Equivalent data to the lca_prefix.php form, but machine-readable.
    """
    if not npa or not nxx:
        return (None, None)
    try:
        r = requests.get(
            "https://www.localcallingguide.com/xmlprefix.php",
            params={"npa": npa, "nxx": nxx}, timeout=10,
        )
        r.raise_for_status()
        root = ET.fromstring(r.text)
        rc_tag = root.find(".//rc")
        lata_tag = root.find(".//lata")
        rc = (rc_tag.text or "").strip() if rc_tag is not None else None
        lata = (lata_tag.text or "").strip() if lata_tag is not None else None
        return (rc or None, lata or None)
    except Exception:
        return (None, None)

def normalize_rc(s: str | None) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", s.strip().upper())

def eng_value_for_rate_center(eng_df: pd.DataFrame, rate_centre: str) -> str:
    """
    Find the row where Engineering column O matches the given Rate Centre (case/space-insensitive),
    and return the value from column N on that same row. If not found, return "".
    (Excel columns: N->index 13, O->index 14 when header=None)
    """
    if eng_df is None or eng_df.empty:
        return ""
    wanted = normalize_rc(rate_centre)
    O_IDX = 14  # column O (0-based)
    N_IDX = 13  # column N (0-based)
    rows = eng_df.shape[0]
    for r in range(rows):
        val_o = normalize_rc(eng_df.iat[r, O_IDX] if O_IDX < eng_df.shape[1] else "")
        if val_o and val_o == wanted:
            val_n = eng_df.iat[r, N_IDX] if N_IDX < eng_df.shape[1] else ""
            return "" if (pd.isna(val_n) or str(val_n).lower() == "nan") else str(val_n)
    return ""

# ---------- Main ----------

if uploaded_file:
    # Read bytes once; reuse for openpyxl + pandas
    file_bytes = uploaded_file.read()
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    try:
        user_details_ws = get_sheet(wb, "User details")
        eng_ws          = get_sheet(wb, "Engineering")
        call_flow_ws    = get_sheet(wb, "Call flow")
    except KeyError as e:
        st.error(str(e)); st.stop()

    st.caption(f"✅ Found sheets: {wb.sheetnames}")

    customer_name = user_details_ws["B3"].value
    region = (eng_ws["C4"].value or "").strip()  # CH or LV

    # BG defaults used in Business Group section (unchanged)
    lcc_defaults = {
        "lcc1":  str(eng_ws["C17"].value or ""),
        "lcc2":  str(eng_ws["C18"].value or ""),
        "lcc3":  str(eng_ws["C12"].value or ""),
        "lcc15": str(eng_ws["C19"].value or ""),
    }

    st.success(f"Loaded file for **{customer_name}** (Region: {region})")

    # DataFrames for flexible parsing
    user_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=user_details_ws.title, header=None)
    eng_df  = pd.read_excel(io.BytesIO(file_bytes), sheet_name=eng_ws.title, header=None, dtype=str)

    # Column indexes (0-based) for User details
    COL_NAME     = 0   # A
    COL_PHONE    = 1   # B
    COL_CALLING  = 3   # D
    COL_EXT      = 4   # E
    COL_EMAIL    = 5   # F
    COL_ACCT     = 7   # H
    COL_DEPT     = 8   # I
    COL_TZ       = 9   # J
    COL_TEMPLATE = 12  # M
    COL_MAC      = 13  # N
    COL_MLHG     = 14  # O
    COL_LCC15    = 17  # R  <-- NEW: Line Class Code 15 per-subscriber

    START_ROW = 8  # Excel row 9

    # =========================
    # Build BG (width=16; includes BG LCC defaults)
    # =========================
    BG_COLS = 16
    def pad_bg(values): return (values + [""] * max(0, BG_COLS - len(values)))[:BG_COLS]
    bg_template = f"{region} BG"

    numbers = []
    for cell in user_details_ws["B9":"B100"]:
        for c in cell:
            if c.value:
                numbers.append(str(c.value).strip())
    for cell in call_flow_ws["D17":"D27"]:
        for c in cell:
            if c.value:
                numbers.append(str(c.value).strip())
    numbers = [n for n in dict.fromkeys(numbers) if n]

    departments, seen_depts = [], set()
    for cell in user_details_ws["I9":"I100"]:
        for c in cell:
            if c.value:
                d = str(c.value).strip()
                if d and d not in seen_depts:
                    seen_depts.add(d); departments.append(d)

    bg_rows = []
    bg_rows.append(pad_bg(["#"]))
    bg_rows.append(pad_bg(["#"]))
    bg_rows.append(pad_bg(["#"]))
    bg_rows.append(pad_bg(["#Business Groups"]))
    bg_rows.append(pad_bg(["Business Group"]))
    bg_rows.append(pad_bg([
        "MetaSphere CFS","MetaSphere EAS","Business Group","Template","CFS Persistent Profile",
        "Local CNAM name","Music On Hold Service - Subscribed","Music On Hold Service - class of service",
        "Music On Hold Service - limit concurrent calls","Music On Hold Service - maximum concurrent calls",
        "Music On Hold Service - Service Level","Music On Hold Service - Application Server",
        "Line Class Code 1","Line Class Code 2","Line Class Code 3","Line Class Code 15",
    ]))
    bg_rows.append(pad_bg([
        "CommandLink","CommandLink_vEAS_LV",customer_name,bg_template,bg_template,"",
        "TRUE","0","", "16","Enhanced","EAS Voicemail",
        lcc_defaults["lcc1"], lcc_defaults["lcc2"], lcc_defaults["lcc3"], lcc_defaults["lcc15"],
    ]))
    bg_rows.append(pad_bg([""]))
    bg_rows.append(pad_bg([""]))
    bg_rows.append(pad_bg(["#BG Number Blocks"]))
    bg_rows.append(pad_bg(["Business Group Number Block"]))
    bg_rows.append(pad_bg([
        "MetaSphere CFS","Business Group","First Phone Number","Block size","CFS Subscriber Group",
    ]))
    for num in numbers:
        bg_rows.append(pad_bg(["CommandLink",customer_name,num,"1","Standard Subscribers"]))
    bg_rows.append(pad_bg([""]))
    bg_rows.append(pad_bg([""]))
    bg_rows.append(pad_bg([""]))
    bg_rows.append(pad_bg(["Department"]))
    bg_rows.append(pad_bg(["MetaSphere CFS","MetaSphere EAS","Business Group","Name"]))
    for dept in departments:
        bg_rows.append(pad_bg(["CommandLink","CommandLink_vEAS_LV",customer_name,dept]))

    # =========================
    # Seats/Devices/Exts/MLHG
    # Subscribers: add 4 new LCC fields after "use local name ..."
    # =========================
    SEATS_COLS = 32  # 28 + 4 new LCC columns
    def pad27(values): return (values + [""] * max(0, SEATS_COLS - len(values)))[:SEATS_COLS]

    sub_rows = []
    sub_rows.append(pad27(["#"]))
    sub_rows.append(pad27(["#"]))
    sub_rows.append(pad27(["#"]))
    sub_rows.append(pad27(["#BG Subscriber"]))
    sub_rows.append(pad27(["Subscriber"]))
    sub_rows.append(pad27([
        "MetaSphere CFS","MetaSphere EAS","Phone number","Template","Business Group (CFS)","Business Group (EAS)",
        "CFS Subscriber Group","Name (CFS)","Name (EAS)","PIN (CFS)","PIN (EAS)","EAS Preferred Language",
        "EAS Customer Group","EAS Password","Business Group Administration - account type (CFS)",
        "Business Group Administration - account type (EAS)","Line State Monitoring - Subscribed",
        "Calling Name Delivery - local name (BG subscriber)","Account Email","Timezone (CFS)","Timezone (EAS)",
        "Calling party number","Charge number","Calling party number for emergency calls","Department (CFS)",
        "Department (EAS)","Calling Name Delivery - use local name for intra-BG calls only",
        # NEW subscriber LCC fields:
        "Line Class Code 1","Line Class Code 2","Line Class Code 3","Line Class Code 15",
    ]))

    for i in range(START_ROW, len(user_df)):
        name         = user_df.iloc[i, COL_NAME]
        phone        = user_df.iloc[i, COL_PHONE]
        calling      = user_df.iloc[i, COL_CALLING]
        email        = user_df.iloc[i, COL_EMAIL]
        account_type = user_df.iloc[i, COL_ACCT]
        department   = user_df.iloc[i, COL_DEPT]
        tz_val       = user_df.iloc[i, COL_TZ]
        template_raw = user_df.iloc[i, COL_TEMPLATE]
        lcc15_cell   = user_df.iloc[i, COL_LCC15]  # R: per-subscriber LCC15

        # Skip blank phone or reserved/none templates
        if pd.isna(phone) or str(template_raw).strip() in ["None", "Reserve Number", "None | Reserve Number"]:
            continue

        # Template + flags
        template = convert_template(template_raw, region)
        is_aa = template in [f"{region}_AA_Easy", f"{region}_AA_Premium"]
        line_state_monitor     = "" if is_aa else "TRUE"
        calling_name_delivery  = "" if is_aa else ("" if pd.isna(name) else str(name))
        intra_bg_calls         = "" if is_aa else "TRUE"
        acct_value             = "Administrator" if str(account_type) in ["Location Admin","Company Admin"] else "Normal"
        tz_cfs = "" if pd.isna(tz_val) else str(tz_val)
        tz_eas = tz_cfs

        # ---- NEW: LCCs per subscriber
        # LCC15 → from column R
        lcc15 = "" if pd.isna(lcc15_cell) else str(lcc15_cell).strip()

        # LCC3 → area code (first 3 digits)
        npa, nxx, _ = npanxx_from_number(phone)
        lcc3 = npa or ""

        # Lookup Rate Centre + LATA by NPA-NXX
        rc, lata = lookup_rc_lata(npa, nxx) if (npa and nxx) else (None, None)
        # LCC2 → LATA
        lcc2 = lata or ""

        # LCC1 → from Engineering: find row where column O == Rate Centre; return column N
        lcc1 = eng_value_for_rate_center(eng_df, rc) if rc else ""

        sub_rows.append(pad27([
            "CommandLink","CommandLink_vEAS_LV",
            str(phone),template,customer_name,customer_name,
            "Standard Subscribers",
            "" if pd.isna(name) else str(name),
            "" if pd.isna(name) else str(name),
            "","",
            "eng","defaultGroup","",
            acct_value,acct_value,
            line_state_monitor,calling_name_delivery,
            "" if pd.isna(email) else str(email),
            tz_cfs,tz_eas,
            "" if pd.isna(calling) else str(calling),
            str(phone),str(phone),
            "" if pd.isna(department) else str(department),
            "" if pd.isna(department) else str(department),
            intra_bg_calls,
            # New fields in the specified order:
            lcc1, lcc2, lcc3, lcc15,
        ]))

    # spacer
    sub_rows.append(pad27([""]))

    # ---- Managed Device ----
    sub_rows.append(pad27(["#Managed Device"]))
    sub_rows.append(pad27(["Managed Device"]))
    sub_rows.append(pad27([
        "MetaSphere CFS","Business Group","MAC address","Assigned to user","User directory number",
        "MAC trusted until","Device version","Device model","Description",
    ]))
    for i in range(START_ROW, len(user_df)):
        phone = user_df.iloc[i, COL_PHONE]
        mac   = user_df.iloc[i, COL_MAC]
        if pd.isna(phone) or pd.isna(mac) or str(mac).strip() == "":
            continue
        sub_rows.append(pad27([
            "CommandLink",customer_name,str(mac),"TRUE",str(phone),
            mac_trusted_until_str(),"2","Determined by Endpoint Pack","",
        ]))

    # ---- Intercom Code Range ----
    sub_rows.append(pad27([""])); sub_rows.append(pad27([""])); sub_rows.append(pad27([""]))
    sub_rows.append(pad27(["#Intercom Code Range"]))
    sub_rows.append(pad27(["Intercom Code Range"]))
    sub_rows.append(pad27([
        "MetaSphere CFS","MetaSphere EAS","Business Group","First Code","Last Code","First Directory Number",
    ]))
    for i in range(START_ROW, len(user_df)):
        phone = user_df.iloc[i, COL_PHONE]
        ext   = user_df.iloc[i, COL_EXT]
        if pd.isna(phone) or pd.isna(ext) or str(ext).strip() == "":
            continue
        sub_rows.append(pad27([
            "CommandLink","CommandLink_vEAS_LV",customer_name,str(ext),str(ext),str(phone),
        ]))

    # ---- MLHGs ----
    sub_rows.append(pad27([""])); sub_rows.append(pad27([""])); sub_rows.append(pad27([""])); sub_rows.append(pad27([""]))
    sub_rows.append(pad27(["#MLHGs"]))
    sub_rows.append(pad27(["MLHG"]))
    sub_rows.append(pad27([
        "MetaSphere CFS","Business Group","MLHG Name",
        "Members;Directory number;Login/logout supported","Distribution algorithm","Hunt on no-answer",
    ]))
    for r in range(17, 28):
        mlg_name = call_flow_ws[f"B{r}"].value
        if not mlg_name:
            continue
        dist_alg = call_flow_ws[f"C{r}"].value
        dist_alg_clean = "" if pd.isna(dist_alg) else str(dist_alg).strip()
        if dist_alg_clean == "Ring All":
            dist_alg_clean = "Ring all"
        members = []
        for i in range(START_ROW, len(user_df)):
            if str(user_df.iloc[i, COL_MLHG]).strip() == str(mlg_name).strip():
                num = user_df.iloc[i, COL_PHONE]
                if pd.notna(num):
                    members.append(f"{{'{str(num)}';'FALSE'}}")
        sub_rows.append(pad27([
            "CommandLink",customer_name,str(mlg_name),";".join(members),dist_alg_clean,"FALSE",
        ]))

    # ---- MLHG Pilot ----
    sub_rows.append(pad27([""])); sub_rows.append(pad27([""]))
    sub_rows.append(pad27(["#MLHG Pilot"]))
    sub_rows.append(pad27(["MLHG Pilot Number"]))
    sub_rows.append(pad27([
        "MetaSphere CFS","MetaSphere EAS","Business Group (CFS)","MLHG Name","Phone number",
        "Template","Name (EAS)","Name (CFS)","PIN (EAS)","EAS Password","EAS Customer Group",
    ]))
    for r in range(17, 28):
        mlg_name     = call_flow_ws[f"B{r}"].value
        phone_number = call_flow_ws[f"D{r}"].value
        pilot_vm     = call_flow_ws[f"H{r}"].value
        if not mlg_name or not phone_number:
            continue
        pilot_template = f"{region}_MLHG_Pilot" if str(pilot_vm).strip().lower() == "yes" else f"{region}_MLHG_Pilot_NoVM"
        sub_rows.append(pad27([
            "CommandLink","CommandLink_vEAS_LV",customer_name,str(mlg_name),str(phone_number),
            pilot_template,f"{mlg_name} Pilot",f"{mlg_name} Pilot","*","*","defaultGroup",
        ]))

    # ---------- Single combined CSV (BG then Seats) ----------
    combined_buffer = io.StringIO()
    writer = csv.writer(combined_buffer, lineterminator="\n")
    writer.writerows(bg_rows)
    writer.writerow([])
    writer.writerows(sub_rows)

    combined_filename = f"{customer_name}-Meta-Import-Combined.csv"
    st.download_button(
        label=f"⬇️ Download {combined_filename}",
        data=combined_buffer.getvalue(),
        file_name=combined_filename,
        mime="text/csv",
    )

else:
    st.info("Please upload an Excel file to begin.")
